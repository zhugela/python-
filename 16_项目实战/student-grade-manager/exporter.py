"""Excel成绩报表导出"""

import os
from file_handler import get_all_students
from analyzer import get_all_subjects


def export_excel(file_path=None):
    """导出成绩报表到Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl 未安装，请运行: pip install openpyxl")
        return False

    students = get_all_students()
    if not students:
        print("暂无学生数据，无法导出")
        return False

    valid = [s for s in students if s.grades]
    if not valid:
        print("暂无成绩数据，无法导出")
        return False

    if file_path is None:
        file_path = os.path.join(os.path.dirname(__file__), "成绩报表.xlsx")

    # 按总分排序
    ranked = sorted(valid, key=lambda s: s.get_total_score(), reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "成绩报表"

    # 样式定义
    title_font = Font(name="微软雅黑", size=16, bold=True)
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    data_font = Font(name="微软雅黑", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 合并标题行
    subjects = get_all_subjects(valid)
    total_cols = 4 + len(subjects)  # 排名 + 学号 + 姓名 + 科目... + 总分 + 平均分

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value="学生成绩报表")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # 表头
    headers = ["排名", "学号", "姓名"] + list(subjects) + ["总分", "平均分"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 30

    # 数据行
    rank_fill_gold = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    rank_fill_silver = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for i, s in enumerate(ranked):
        row = i + 3

        # 排名（前3名特殊标记）
        rank_cell = ws.cell(row=row, column=1, value=i + 1)
        rank_cell.alignment = center_align
        rank_cell.border = thin_border
        rank_cell.font = Font(name="微软雅黑", size=11, bold=(i < 3))
        if i == 0:
            rank_cell.fill = rank_fill_gold
        elif i == 1:
            rank_cell.fill = rank_fill_silver

        # 学号
        cell = ws.cell(row=row, column=2, value=s.student_id)
        cell.font = data_font
        cell.alignment = center_align
        cell.border = thin_border

        # 姓名
        cell = ws.cell(row=row, column=3, value=s.name)
        cell.font = data_font
        cell.alignment = center_align
        cell.border = thin_border

        # 各科成绩
        for j, sub in enumerate(subjects):
            cell = ws.cell(row=row, column=4 + j, value=s.grades.get(sub, ""))
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            # 不及格标红
            if isinstance(cell.value, (int, float)) and cell.value < 60:
                cell.font = Font(name="微软雅黑", size=11, color="FF0000", bold=True)

        # 总分
        cell = ws.cell(row=row, column=4 + len(subjects), value=s.get_total_score())
        cell.font = Font(name="微软雅黑", size=11, bold=True)
        cell.alignment = center_align
        cell.border = thin_border

        # 平均分
        cell = ws.cell(row=row, column=5 + len(subjects), value=s.get_average_score())
        cell.font = Font(name="微软雅黑", size=11, bold=True)
        cell.alignment = center_align
        cell.border = thin_border

    # 列宽
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    for j in range(len(subjects)):
        ws.column_dimensions[chr(68 + j)].width = 12
    ws.column_dimensions[chr(68 + len(subjects))].width = 10
    ws.column_dimensions[chr(69 + len(subjects))].width = 10

    wb.save(file_path)
    print(f"✅ 成绩报表已导出到: {file_path}")
    return True
